"""
Rutas API para el Asistente de Facturación SAT
"""
from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from pydantic import BaseModel
from typing import List, Dict, Any, Optional
import tempfile
import os

from services.asistente_facturacion_service import asistente_facturacion_service

router = APIRouter(prefix="/asistente-facturacion", tags=["Asistente Facturación"])

MAX_FILE_SIZE = 10 * 1024 * 1024


class ChatRequest(BaseModel):
    mensaje: str
    historial: Optional[List[Dict[str, Any]]] = None


class ChatResponse(BaseModel):
    respuesta: str
    sugerencias: List[Dict[str, Any]]


@router.post("/chat", response_model=ChatResponse)
async def chat_facturacion(request: ChatRequest):
    """
    Chat con el asistente de facturación SAT
    Recibe un mensaje y retorna respuesta con sugerencias de claves SAT
    """
    if not request.mensaje or not request.mensaje.strip():
        raise HTTPException(status_code=400, detail="Mensaje requerido")
    
    result = await asistente_facturacion_service.chat(
        mensaje=request.mensaje,
        historial=request.historial or []
    )
    
    return ChatResponse(
        respuesta=result.get("respuesta", ""),
        sugerencias=result.get("sugerencias", [])
    )


@router.post("/analizar")
async def analizar_documento(
    file: UploadFile = File(...),
    accion: str = Form(default="analizar_facturacion")
):
    """
    Analiza un documento para sugerir claves SAT
    Soporta PDF, DOCX, DOC, TXT, XLSX (máximo 10MB)
    """
    if not file:
        raise HTTPException(status_code=400, detail="Archivo requerido")
    
    allowed_extensions = {'.pdf', '.doc', '.docx', '.txt', '.xlsx', '.xls'}
    filename = file.filename or "documento"
    ext = os.path.splitext(filename)[1].lower()
    
    if ext not in allowed_extensions:
        raise HTTPException(
            status_code=400, 
            detail=f"Tipo de archivo no soportado: {ext}. Usa: {', '.join(allowed_extensions)}"
        )
    
    try:
        content = await file.read()
        
        if len(content) > MAX_FILE_SIZE:
            raise HTTPException(
                status_code=400,
                detail=f"Archivo demasiado grande. Máximo permitido: 10MB"
            )
        texto_extraido = await _extraer_texto(content, ext, filename)
        
        if not texto_extraido or len(texto_extraido.strip()) < 50:
            raise HTTPException(
                status_code=400,
                detail="No se pudo extraer suficiente texto del documento"
            )
        
        result = await asistente_facturacion_service.analizar_documento(
            contenido_texto=texto_extraido,
            nombre_archivo=filename
        )
        
        return {
            "analisis": result.get("analisis", ""),
            "sugerencias": result.get("sugerencias", []),
            "documento": result.get("documento", {})
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error analizando documento: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/voz")
async def procesar_voz(
    audio: UploadFile = File(...)
):
    """
    Procesa un mensaje de voz
    TODO: Integrar con servicio de transcripción (Whisper/ElevenLabs)
    """
    if not audio:
        raise HTTPException(status_code=400, detail="Audio requerido")
    
    transcripcion = "[Función de transcripción de voz próximamente disponible]"
    
    result = await asistente_facturacion_service.chat(
        mensaje="Necesito ayuda para facturar un servicio",
        historial=[]
    )
    
    return {
        "transcripcion": transcripcion,
        "respuesta": result.get("respuesta", ""),
        "sugerencias": result.get("sugerencias", [])
    }


async def _extraer_texto(content: bytes, extension: str, filename: str) -> str:
    """Extrae texto de diferentes tipos de documentos"""
    
    if extension == '.txt':
        try:
            return content.decode('utf-8')
        except:
            return content.decode('latin-1')
    
    elif extension == '.pdf':
        try:
            import fitz
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
                tmp.write(content)
                tmp_path = tmp.name
            
            doc = fitz.open(tmp_path)
            texto = ""
            for page in doc:
                texto += page.get_text()
            doc.close()
            os.unlink(tmp_path)
            return texto
        except ImportError:
            try:
                import pypdf
                import io
                reader = pypdf.PdfReader(io.BytesIO(content))
                texto = ""
                for page in reader.pages:
                    texto += page.extract_text() or ""
                return texto
            except ImportError:
                return f"[Documento PDF: {filename} - Instalar PyMuPDF o pypdf para extraer texto]"
        except Exception as e:
            return f"[Error extrayendo texto del PDF: {str(e)}]"
    
    elif extension in ['.docx', '.doc']:
        try:
            import docx
            import io
            doc = docx.Document(io.BytesIO(content))
            texto = "\n".join([para.text for para in doc.paragraphs])
            return texto
        except ImportError:
            return f"[Documento Word: {filename} - Instalar python-docx para extraer texto]"
        except Exception as e:
            return f"[Error extrayendo texto del documento Word: {str(e)}]"
    
    elif extension in ['.xlsx', '.xls']:
        try:
            import openpyxl
            import io
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            texto = ""
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) for cell in row if cell is not None])
                    if row_text:
                        texto += row_text + "\n"
            return texto
        except ImportError:
            return f"[Archivo Excel: {filename} - Instalar openpyxl para extraer texto]"
        except Exception as e:
            return f"[Error extrayendo texto del Excel: {str(e)}]"
    
    return f"[Tipo de archivo no procesable: {extension}]"
